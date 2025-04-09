import pymysql
from pymysql.cursors import DictCursor
import openpyxl


def get_connection():
    return pymysql.connect(
        host="localhost",
        user="team1",
        password="team1",
        database="videojuego",
        cursorclass=DictCursor,
    )


def cargar_maestro(archivo_xlsx):
    try:
        wb = openpyxl.load_workbook(archivo_xlsx)
        sheet = wb.active

        for row in sheet.iter_rows(min_row=2, values_only=True):
            Correo, Grupo, Nombre = row

            connection = get_connection()
            with connection.cursor() as cursor:
                # revisa si ya hay esos elementos para evitar duplicados
                if cursor.execute(
                    "SELECT * FROM Maestro WHERE Correo = %s AND Grupo = %s AND Nombre = %s",
                    (Correo, Grupo, Nombre),
                ):
                    continue
                else:
                    cursor.execute(
                        "INSERT INTO Maestro (Correo, Grupo, Nombre) VALUES (%s, %s, %s)",
                        (Correo, Grupo, Nombre),
                    )
            connection.commit()
        print(f"Archivo {archivo_xlsx} cargado correctamente.")
    except Exception as e:
        print(f"Error al cargar {archivo_xlsx}: {e}")


def cargar_alumno(archivo_xlsx):
    try:
        wb = openpyxl.load_workbook(archivo_xlsx)
        sheet = wb.active

        for row in sheet.iter_rows(min_row=2, values_only=True):
            NumLista, Genero, Grupo, IDMaestro = row

            connection = get_connection()
            with connection.cursor() as cursor:
                if cursor.execute(
                    "SELECT * FROM Alumno WHERE NumLista = %s AND Genero = %s AND Grupo = %s AND IDMaestro = %s",
                    (NumLista, Genero, Grupo, IDMaestro),
                ):
                    continue
                else:
                    cursor.execute(
                        "INSERT INTO Alumno (Genero, Grupo, NumLista, IDMaestro) VALUES (%s, %s, %s, %s)",
                        (Genero, Grupo, NumLista, IDMaestro),
                    )
            connection.commit()
        print(f"Archivo {archivo_xlsx} cargado correctamente.")
    except Exception as e:
        print(f"Error al cargar {archivo_xlsx}: {e}")


def cargar_nivel(archivo_xlsx):
    try:
        wb = openpyxl.load_workbook(archivo_xlsx)
        sheet = wb.active

        for row in sheet.iter_rows(min_row=2, values_only=True):
            NumNivel, NumPreguntas, Descripcion = row

            connection = get_connection()
            with connection.cursor() as cursor:
                if cursor.execute(
                    "SELECT * FROM Nivel WHERE NumNivel = %s AND NumPreguntas = %s AND Descripcion = %s",
                    (NumNivel, NumPreguntas, Descripcion),
                ):
                    continue
                else:
                    cursor.execute(
                        "INSERT INTO Nivel (NumNivel, NumPreguntas, Descripcion) VALUES (%s, %s, %s)",
                        (NumNivel, NumPreguntas, Descripcion),
                    )
            connection.commit()
        print(f"Archivo {archivo_xlsx} cargado correctamente.")
    except Exception as e:
        print(f"Error al cargar {archivo_xlsx}: {e}")


def cargar_pregunta(archivo_xlsx):
    try:
        wb = openpyxl.load_workbook(archivo_xlsx)
        sheet = wb.active

        for row in sheet.iter_rows(min_row=2, values_only=True):
            Texto, NumNivel, Respuesta = row

            connection = get_connection()
            with connection.cursor() as cursor:
                if cursor.execute(
                    "SELECT * FROM Pregunta WHERE Texto = %s AND NumNivel = %s AND Respuesta = %s",
                    (Texto, NumNivel, Respuesta),
                ):
                    continue
                else:
                    cursor.execute(
                        "INSERT INTO Pregunta (Texto, NumNivel, Respuesta) VALUES (%s, %s, %s)",
                        (Texto, NumNivel, Respuesta),
                    )
            connection.commit()
        print(f"Archivo {archivo_xlsx} cargado correctamente.")
    except Exception as e:
        print(f"Error al cargar {archivo_xlsx}: {e}")


cargar_maestro("RegistroMaestros.xlsx")
cargar_alumno("RegistroAlumnos.xlsx")
cargar_nivel("RegistroNivel.xlsx")
cargar_pregunta("RegistroPregunta.xlsx")

